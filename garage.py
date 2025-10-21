# ================================================================
# APP STREAMLIT - Générateur d'écritures comptables pour Pennylane
# ================================================================

import re
from datetime import datetime
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ================================================================
# AUTHENTIFICATION
# ================================================================
if "login" not in st.session_state:
    st.session_state["login"] = False
if "page" not in st.session_state:
    st.session_state["page"] = "Accueil"  # page par défaut

def login(username, password):
    users = {
        "aurore": {"password": "12345", "name": "Aurore Demoulin"},
        "laure.froidefond": {"password": "Laure2019$", "name": "Laure Froidefond"},
        "bruno": {"password": "Toto1963$", "name": "Toto El Gringo"}
    }
    if username in users and password == users[username]["password"]:
        st.session_state["login"] = True
        st.session_state["username"] = username
        st.session_state["name"] = users[username]["name"]
        st.session_state["page"] = "Accueil"
        st.success(f"Bienvenue {st.session_state['name']} 👋")
        st.rerun()
    else:
        st.error("❌ Identifiants incorrects")

if not st.session_state["login"]:
    st.title("🔑 Connexion espace expert-comptable")
    username_input = st.text_input("Identifiant")
    password_input = st.text_input("Mot de passe", type="password")
    if st.button("Connexion"):
        login(username_input, password_input)
    st.stop()

# ================================================================
# OUTIL DE GENERATION EXCEL
# ================================================================

def generer_code_client(nom_client: str) -> str:
    """Génère un code client du type 411X00000 à partir du nom."""
    nom_client = nom_client.strip().upper()
    premiere_lettre = nom_client[0] if nom_client else "X"
    return f"411{premiere_lettre}00000"

def generer_ecritures_excel(fichier_txt):
    """Lit le fichier texte et renvoie un fichier Excel binaire (BytesIO)."""
    lignes = [line.strip().split(',') for line in fichier_txt.decode('utf-8').splitlines() if line.strip()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ecritures"

    entetes = ["Date", "Journal", "Numéro de compte", "Libellé", "Montant au débit", "Montant au crédit"]
    ws.append(entetes)

    for ligne in lignes:
        if len(ligne) < 11:
            continue
        
        _, date_raw, journal, _, _, libelle_brut, facture, montant_raw, sens, _, devise = ligne

        try:
            date = datetime.strptime(date_raw, "%d%m%y").strftime("%d/%m/%Y")
        except ValueError:
            continue
        
        montant = round(float(montant_raw), 2)

        # Extraction du nom client
        match = re.search(r'Fact:\d+\s+([A-ZÉÈA-Z\- ]+)', libelle_brut.upper())
        nom_client = match.group(1).strip().title() if match else "Client Inconnu"

        compte_client = generer_code_client(nom_client)
        libelle = f"Facture {facture} - {nom_client}"

        # Ligne client (débit)
        ws.append([date, journal, compte_client, libelle, montant if sens == "D" else 0, montant if sens == "C" else 0])
        # Ligne vente (crédit)
        ws.append([date, journal, "706000", libelle, montant if sens == "C" else 0, montant if sens == "D" else 0])

    # Ajustement largeur colonnes
    for col_idx, _ in enumerate(entetes, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Sauvegarde en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ================================================================
# INTERFACE UTILISATEUR
# ================================================================
st.title("📘 Générateur d’écritures Pennylane")
st.write(f"Connecté en tant que **{st.session_state['name']}**")

st.divider()

uploaded_file = st.file_uploader("📂 Importer le fichier ventes (.txt)", type=["txt"])

if uploaded_file:
    st.success("✅ Fichier importé avec succès. Cliquez sur le bouton ci-dessous pour générer l’export Excel.")

    if st.button("🚀 Générer le fichier Excel"):
        fichier_excel = generer_ecritures_excel(uploaded_file.read())
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=fichier_excel,
            file_name="ecritures_pennylane.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("✅ Export prêt au téléchargement.")